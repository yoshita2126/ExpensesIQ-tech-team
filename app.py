from flask import Flask, render_template, render_template_string, request, redirect
from flask import session, url_for, jsonify, Response

import sqlite3
import re
import secrets
import io
import csv
import json
from pathlib import Path

import os
import jwt
from services.ai_service import generate_insights
from services.email_service import send_email
from services.forecast_service import forecast_monthly_expense
from datetime import date, datetime, timedelta, timezone
import calendar
from functools import wraps
from typing import Any, Callable, Dict, List, Optional, cast

from werkzeug.security import (
    generate_password_hash,
    check_password_hash
)

app = Flask(__name__)

app.secret_key = secrets.token_hex(32)

# Register API routes
from api_routes import register_api_routes
register_api_routes(app)


# =========================================================
# VALIDATIONS
# =========================================================

def valid_username(username: str) -> bool:

    return len(username) >= 3


def valid_password(password: str) -> bool:

    pattern = (
        r"^(?=.*[a-z])"
        r"(?=.*[A-Z])"
        r"(?=.*\d)"
        r"(?=.*[@$!%*?&])"
        r".{8,}$"
    )

    return bool(re.match(pattern, password))


def valid_gmail(email: str) -> bool:

    pattern = r'^[a-zA-Z0-9._%+-]+@gmail\.com$'

    return bool(re.match(pattern, email))


def valid_amount(amount: Any) -> bool:

    try:

        value = float(amount)

        return value > 0

    except Exception:

        return False


# =========================================================
# CORS
# =========================================================

@app.after_request
def add_cors(response: Response) -> Response:

    response.headers["Access-Control-Allow-Origin"] = "*"

    response.headers["Access-Control-Allow-Headers"] = (
        "Content-Type,Authorization"
    )

    response.headers["Access-Control-Allow-Methods"] = (
        "GET,POST,OPTIONS"
    )

    return response


def get_json_data() -> Dict[str, Any]:
    raw_data = request.get_json(silent=True)
    return raw_data if isinstance(raw_data, dict) else {}


# =========================================================
# DATABASE
# =========================================================

def get_db():

    db_path = Path(__file__).resolve().parent / "data.db"
    conn = sqlite3.connect(str(db_path), timeout=10, check_same_thread=False)

    conn.row_factory = sqlite3.Row

    return conn


def init_db():

    conn = get_db()

    cur = conn.cursor()

    # USERS
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users(

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            username TEXT UNIQUE NOT NULL,

            email TEXT UNIQUE,

            phone TEXT,

            password TEXT NOT NULL,

            role TEXT,

            group_name TEXT,

            registration_date TEXT,

            daily_limit REAL DEFAULT 1000,
  
            email_verified INTEGER DEFAULT 0
        )
    """)

    # Add monthly_budget column to users if it doesn't exist
    cur.execute("PRAGMA table_info(users)")
    cols = [r[1] for r in cur.fetchall()]
    if 'monthly_budget' not in cols:
        try:
            cur.execute("ALTER TABLE users ADD COLUMN monthly_budget REAL DEFAULT NULL")
        except Exception:
            # ignore if alter fails for any reason
            pass

    # GROUPS
    cur.execute("""
        CREATE TABLE IF NOT EXISTS groups(

            group_name TEXT PRIMARY KEY,

            password TEXT NOT NULL,

            role TEXT,

            registration_date TEXT
        )
    """)

    # TRANSACTIONS
    cur.execute("""
        CREATE TABLE IF NOT EXISTS transactions(

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            group_name TEXT,

            user TEXT,

            amount REAL,

            category TEXT,

            type TEXT,

            note TEXT,

            date TEXT
        )
    """)

    # GOALS
    cur.execute("""
        CREATE TABLE IF NOT EXISTS goals(

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            group_name TEXT,

            user TEXT,

            title TEXT,

            target REAL,

            saved REAL,

            due_date TEXT
        )
    """)

    # PAYMENTS
    cur.execute("""
        CREATE TABLE IF NOT EXISTS payments(

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            group_name TEXT,

            user TEXT,

            amount REAL,

            direction TEXT,

            description TEXT,

            status TEXT,

            due_date TEXT
        )
    """)

    # RECURRING PAYMENTS
    cur.execute("""
        CREATE TABLE IF NOT EXISTS recurring_payments(

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            group_name TEXT,

            user TEXT,

            amount REAL,

            category TEXT,

            type TEXT,

            note TEXT,

            start_date TEXT,

            frequency_days INTEGER,

            next_date TEXT,

            active INTEGER DEFAULT 1
        )
    """)

    # OTP TOKENS
    cur.execute("""
        CREATE TABLE IF NOT EXISTS otp_tokens(

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            email TEXT,

            phone TEXT,

            otp TEXT NOT NULL,

            expires_at TEXT NOT NULL,

            created_at TEXT NOT NULL
        )
    """)

    # BUSINESSES
    cur.execute("""
        CREATE TABLE IF NOT EXISTS businesses(

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            business_name TEXT NOT NULL,

            owner_id INTEGER NOT NULL,

            gst_id TEXT UNIQUE,

            pan TEXT,

            address TEXT,

            city TEXT,

            state TEXT,

            pincode TEXT,

            phone TEXT,

            email TEXT,

            business_type TEXT,

            industry TEXT,

            registration_date TEXT,

            status TEXT DEFAULT 'active',

            FOREIGN KEY(owner_id) REFERENCES users(id)
        )
    """)

    # BRANCHES
    cur.execute("""
        CREATE TABLE IF NOT EXISTS branches(

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            business_id INTEGER NOT NULL,

            branch_name TEXT NOT NULL,

            address TEXT,

            city TEXT,

            state TEXT,

            pincode TEXT,

            phone TEXT,

            manager_id INTEGER,

            registration_date TEXT,

            status TEXT DEFAULT 'active',

            FOREIGN KEY(business_id) REFERENCES businesses(id),

            FOREIGN KEY(manager_id) REFERENCES users(id)
        )
    """)

    # PRODUCTS
    cur.execute("""
        CREATE TABLE IF NOT EXISTS products(

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            business_id INTEGER NOT NULL,

            product_name TEXT NOT NULL,

            sku TEXT UNIQUE,

            category TEXT,

            unit_price REAL,

            cost_price REAL,

            tax_rate REAL DEFAULT 0,

            description TEXT,

            registration_date TEXT,

            status TEXT DEFAULT 'active',

            FOREIGN KEY(business_id) REFERENCES businesses(id)
        )
    """)

    # INVENTORY
    cur.execute("""
        CREATE TABLE IF NOT EXISTS inventory(

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            branch_id INTEGER NOT NULL,

            product_id INTEGER NOT NULL,

            stock_quantity REAL DEFAULT 0,

            reorder_level REAL DEFAULT 10,

            last_updated TEXT,

            FOREIGN KEY(branch_id) REFERENCES branches(id),

            FOREIGN KEY(product_id) REFERENCES products(id)
        )
    """)

    # STOCK TRANSACTIONS
    cur.execute("""
        CREATE TABLE IF NOT EXISTS stock_transactions(

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            inventory_id INTEGER NOT NULL,

            transaction_type TEXT,

            quantity REAL,

            reference_id TEXT,

            notes TEXT,

            transaction_date TEXT,

            FOREIGN KEY(inventory_id) REFERENCES inventory(id)
        )
    """)

    # SUPPLIERS
    cur.execute("""
        CREATE TABLE IF NOT EXISTS suppliers(

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            business_id INTEGER NOT NULL,

            supplier_name TEXT NOT NULL,

            contact_person TEXT,

            email TEXT,

            phone TEXT,

            address TEXT,

            city TEXT,

            state TEXT,

            pincode TEXT,

            gst_id TEXT,

            payment_terms TEXT,

            registration_date TEXT,

            status TEXT DEFAULT 'active',

            FOREIGN KEY(business_id) REFERENCES businesses(id)
        )
    """)

    # INVOICES (Sales)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS invoices(

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            business_id INTEGER NOT NULL,

            branch_id INTEGER NOT NULL,

            invoice_number TEXT UNIQUE NOT NULL,

            customer_name TEXT NOT NULL,

            customer_email TEXT,

            customer_phone TEXT,

            customer_address TEXT,

            invoice_date TEXT,

            due_date TEXT,

            subtotal REAL DEFAULT 0,

            tax_amount REAL DEFAULT 0,

            total_amount REAL DEFAULT 0,

            payment_status TEXT DEFAULT 'unpaid',

            payment_date TEXT,

            notes TEXT,

            created_by INTEGER,

            created_at TEXT,

            FOREIGN KEY(business_id) REFERENCES businesses(id),

            FOREIGN KEY(branch_id) REFERENCES branches(id),

            FOREIGN KEY(created_by) REFERENCES users(id)
        )
    """)

    # INVOICE ITEMS
    cur.execute("""
        CREATE TABLE IF NOT EXISTS invoice_items(

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            invoice_id INTEGER NOT NULL,

            product_id INTEGER,

            description TEXT NOT NULL,

            quantity REAL,

            unit_price REAL,

            tax_rate REAL DEFAULT 0,

            line_total REAL,

            FOREIGN KEY(invoice_id) REFERENCES invoices(id),

            FOREIGN KEY(product_id) REFERENCES products(id)
        )
    """)

    # PURCHASE ORDERS
    cur.execute("""
        CREATE TABLE IF NOT EXISTS purchase_orders(

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            business_id INTEGER NOT NULL,

            supplier_id INTEGER NOT NULL,

            po_number TEXT UNIQUE NOT NULL,

            order_date TEXT,

            expected_delivery_date TEXT,

            subtotal REAL DEFAULT 0,

            tax_amount REAL DEFAULT 0,

            total_amount REAL DEFAULT 0,

            status TEXT DEFAULT 'pending',

            delivery_date TEXT,

            notes TEXT,

            created_by INTEGER,

            created_at TEXT,

            FOREIGN KEY(business_id) REFERENCES businesses(id),

            FOREIGN KEY(supplier_id) REFERENCES suppliers(id),

            FOREIGN KEY(created_by) REFERENCES users(id)
        )
    """)

    # PO ITEMS
    cur.execute("""
        CREATE TABLE IF NOT EXISTS po_items(

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            po_id INTEGER NOT NULL,

            product_id INTEGER,

            description TEXT NOT NULL,

            quantity REAL,

            unit_price REAL,

            tax_rate REAL DEFAULT 0,

            line_total REAL,

            FOREIGN KEY(po_id) REFERENCES purchase_orders(id),

            FOREIGN KEY(product_id) REFERENCES products(id)
        )
    """)

    # GST TRANSACTIONS
    cur.execute("""
        CREATE TABLE IF NOT EXISTS gst_transactions(

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            business_id INTEGER NOT NULL,

            transaction_type TEXT,

            invoice_id INTEGER,

            po_id INTEGER,

            taxable_amount REAL,

            cgst_rate REAL,

            cgst_amount REAL,

            sgst_rate REAL,

            sgst_amount REAL,

            igst_rate REAL,

            igst_amount REAL,

            total_tax REAL,

            transaction_date TEXT,

            FOREIGN KEY(business_id) REFERENCES businesses(id),

            FOREIGN KEY(invoice_id) REFERENCES invoices(id),

            FOREIGN KEY(po_id) REFERENCES purchase_orders(id)
        )
    """)

    conn.commit()

    conn.close()


init_db()

# JWT settings
JWT_SECRET = os.environ.get('JWT_SECRET') or secrets.token_hex(32)
JWT_ALGO = 'HS256'


def create_jwt(identity: str, group_name: str, role: str, expires_minutes: int = 60*24) -> str:
    now = datetime.now(timezone.utc)
    payload: Dict[str, Any] = {
        'sub': identity,
        'group': group_name,
        'role': role,
        'iat': now,
        'exp': now + timedelta(minutes=expires_minutes)
    }
    token = jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGO)
    return token


def decode_jwt(token: str) -> Optional[Dict[str, Any]]:
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGO])
        return payload
    except jwt.ExpiredSignatureError:
        return None
    except Exception:
        return None


def jwt_required(f: Callable[..., Any]) -> Callable[..., Any]:
    @wraps(f)
    def decorated(*args: Any, **kwargs: Any) -> Any:
        # Check Authorization header first
        auth = request.headers.get('Authorization', '')
        if auth.startswith('Bearer '):
            token = auth.split(' ', 1)[1].strip()
            payload = decode_jwt(token)
            if payload:
                cast(Any, request).jwt_payload = payload
                return f(*args, **kwargs)
            return jsonify(success=False, error='Invalid or expired token'), 401

        # fallback to session cookie
        if 'user' in session:
            return f(*args, **kwargs)

        return jsonify(success=False, error='Authentication required'), 401

    return decorated


# =========================================================
# REGISTER
# =========================================================

@app.route("/", methods=["GET", "POST"])
def register():

    if request.method == "POST":

        username = request.form["username"].strip()

        email = request.form.get("email", "").strip()

        password = request.form["password"].strip()

        confirm_password = request.form.get("confirm_password", "").strip()

        role = request.form["role"]

        group_name = (
            request.form["group_name"].strip()
            or username
        )

        # EMAIL VALIDATION
        if not email or not valid_gmail(email):
            return """
            <h2>Invalid Email</h2>

            <p>
            Please provide a valid Gmail address (e.g., user@gmail.com)
            </p>

            <a href='/'>Go Back</a>
            """

        # USERNAME VALIDATION
        if not valid_username(username):

            return """
            <h2>Invalid Username</h2>

            <p>
            Username must contain at least 3 characters
            </p>

            <a href='/'>Go Back</a>
            """

        # PASSWORD VALIDATION
        if not valid_password(password):

            return """
            <h2>Password Invalid</h2>

            <h3>Password must contain:</h3>

            <ul>
                <li>1 Capital Letter</li>
                <li>1 Small Letter</li>
                <li>1 Number</li>
                <li>1 Special Character</li>
                <li>Minimum 8 Characters</li>
            </ul>

            <a href="/">Go Back</a>
            """

        if password != confirm_password:

            return """
            <h2>Password Mismatch</h2>

            <p>
            Password and confirm password must match.
            </p>

            <a href="/">Go Back</a>
            """

        # HASH PASSWORD
        hashed_password = generate_password_hash(password)

        registration_date = date.today().isoformat()

        conn = get_db()

        cur = conn.cursor()

        try:

            # INSERT USER
            cur.execute(
                """
                INSERT INTO users(

                    username,
                    email,
                    password,
                    role,
                    group_name,
                    registration_date

                )
                VALUES(?,?,?,?,?,?)
                """,
                (
                    username,
                    email,
                    hashed_password,
                    role,
                    group_name,
                    registration_date
                ),
            )

            conn.commit()

        except sqlite3.IntegrityError as e:

            conn.close()

            if "email" in str(e):
                return """
                <h2>Email already registered</h2>

                <a href='/'>Go Back</a>
                """
            else:
                return """
                <h2>Username already exists</h2>

                <a href='/'>Go Back</a>
                """

        conn.close()

        return redirect(url_for("login"))

    return render_template("register.html")


# =========================================================
# LOGIN
# =========================================================

@app.route("/login", methods=["GET", "POST"])
def login():

    if request.method == "POST":

        username = request.form["username"].strip()

        password = request.form["password"].strip()

        conn = get_db()

        cur = conn.cursor()

        # GET USER (case-insensitive lookup by username or email)
        cur.execute(
            "SELECT * FROM users WHERE LOWER(username)=LOWER(?) OR LOWER(email)=LOWER(?)",
            (username, username)
        )

        user = cur.fetchone()

        # CHECK USER PASSWORD
        if user and check_password_hash(
            user["password"],
            password
        ):

            session["user"] = user["username"]

            session["role"] = user["role"]

            session["group_name"] = user["group_name"]

            session["registration_date"] = (
                user["registration_date"]
            )

            session["daily_limit"] = user["daily_limit"] if "daily_limit" in user.keys() else 1000

            conn.close()

            return redirect(url_for("dashboard"))

        conn.close()

        return """
        <h2>Invalid Username, Email, or Password</h2>

        <a href='/login'>Try Again</a>
        """

    return render_template("login.html")


# =========================================================
# LOGOUT
# =========================================================

@app.route("/logout")
def logout():

    session.clear()

    return redirect(url_for("login"))


# =========================================================
# FORGOT PASSWORD
# =========================================================

@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():

    if request.method == "POST":

        account = request.form.get("email", "").strip()
        session.pop("password_reset_email", None)
        session.pop("pending_reset_email", None)

        if not account:
            return render_template("forgot_password.html", error="Username or email is required")

        conn = get_db()
        cur = conn.cursor()

        # Check account and always send to the registered email address.
        cur.execute("SELECT * FROM users WHERE email = ? OR username = ?", (account, account))
        user = cur.fetchone()

        conn.close()

        if not user or not user["email"]:
            return render_template("forgot_password.html", error="Registered email not found")

        email = user["email"]

        # Generate and send OTP
        from services import otp_service
        from services.otp_service import generate_otp, send_otp, store_otp

        otp = generate_otp()

        store_result = store_otp(email=email, otp=otp)

        if not store_result:
            return render_template("forgot_password.html", error="Failed to generate OTP. Try again.")

        send_result = send_otp(email=email, otp=otp)

        if not send_result.get('success'):
            error_message = send_result.get('error') or 'Unknown email send error'
            return render_template("forgot_password.html", error=f"Failed to send OTP: {error_message}")

        session["pending_reset_email"] = email

        return redirect(url_for("reset_password", email=email, sent=1))

    return render_template("forgot_password.html")


@app.route("/verify-otp", methods=["GET", "POST"])
def verify_otp():

    email = request.args.get("email", "").strip()

    if request.method == "POST":

        otp = request.form.get("otp", "").strip()

        if not otp:
            return render_template("verify_otp.html", email=email, error="OTP is required")

        from services.otp_service import verify_otp as verify_otp_func

        if not verify_otp_func(otp=otp, email=email):
            return render_template("verify_otp.html", email=email, error="Invalid or expired OTP")

        session["password_reset_email"] = email

        return redirect(url_for("reset_password", email=email))

    return render_template("verify_otp.html", email=email)


@app.route("/reset-password", methods=["GET", "POST"])
def reset_password():

    email = request.args.get("email", "").strip()
    verified_email = session.get("password_reset_email", "")
    pending_email = session.get("pending_reset_email", "")

    if not email or email not in {verified_email, pending_email}:
        return redirect(url_for("forgot_password"))

    if request.method == "POST":

        otp = request.form.get("otp", "").strip()

        new_password = request.form.get("password", "").strip()

        confirm_password = request.form.get("confirm_password", "").strip()

        from services.otp_service import verify_otp as verify_otp_func

        if not verify_otp_func(otp=otp, email=email):
            return render_template(
                "reset_password.html",
                email=email,
                otp_sent=True,
                error="Invalid or expired OTP"
            )

        if new_password != confirm_password:
            return render_template("reset_password.html", email=email, otp_sent=True, error="Passwords do not match")

        if not valid_password(new_password):
            return render_template("reset_password.html", email=email, otp_sent=True, error="Password must contain: uppercase, lowercase, number, special char, minimum 8 characters")

        hashed_password = generate_password_hash(new_password)

        conn = get_db()
        cur = conn.cursor()

        try:
            cur.execute("UPDATE users SET password = ? WHERE email = ?", (hashed_password, email))
            conn.commit()

            from services.otp_service import delete_otp
            delete_otp(email)
            session.pop("password_reset_email", None)
            session.pop("pending_reset_email", None)

        except Exception as e:
            return render_template("reset_password.html", email=email, otp_sent=True, error=f"Error resetting password: {str(e)}")

        finally:
            conn.close()

        return render_template("reset_success.html")

    return render_template("reset_password.html", email=email, otp_sent=request.args.get("sent") == "1")


# =========================================================
# BUSINESS MODULE
# =========================================================

@app.route("/business/register", methods=["GET", "POST"])
def business_register():
    """Register a new business account"""
    
    if not session.get('user'):
        return redirect(url_for('login'))
    
    if request.method == "POST":
        from services.business_service import register_business
        
        conn = get_db()
        cur = conn.cursor()
        
        # Get user ID
        cur.execute("SELECT id FROM users WHERE username = ?", (session['user'],))
        user = cur.fetchone()
        conn.close()
        
        if not user:
            return render_template("business_register.html", error="User not found")
        
        business_name = request.form.get("business_name", "").strip()
        gst_id = request.form.get("gst_id", "").strip()
        business_type = request.form.get("business_type", "").strip()
        phone = request.form.get("phone", "").strip()
        email = request.form.get("email", "").strip()
        address = request.form.get("address", "").strip()
        city = request.form.get("city", "").strip()
        state = request.form.get("state", "").strip()
        pincode = request.form.get("pincode", "").strip()
        
        if not business_name or not gst_id:
            return render_template("business_register.html", error="Business name and GST ID are required")
        
        result = register_business(
            owner_id=user['id'],
            business_name=business_name,
            gst_id=gst_id,
            business_type=business_type,
            phone=phone,
            email=email,
            address=address,
            city=city,
            state=state,
            pincode=pincode
        )
        
        if result['success']:
            session['business_id'] = result['business_id']
            return redirect(url_for('business_dashboard'))
        else:
            return render_template("business_register.html", error=result.get('error', 'Registration failed'))
    
    return render_template("business_register.html")


@app.route("/business/dashboard")
def business_dashboard():
    """Business dashboard"""
    
    if not session.get('user') or not session.get('business_id'):
        return redirect(url_for('login'))
    
    from services.business_service import (
        get_business_by_id,
        get_business_branches,
        get_business_stats,
        get_business_invoices,
        get_business_purchase_orders,
    )
    
    business_id = session.get('business_id')
    business = get_business_by_id(business_id)
    branches = get_business_branches(business_id)
    stats = get_business_stats(business_id)
    invoices = get_business_invoices(business_id)[:5]
    purchase_orders = get_business_purchase_orders(business_id)[:5]
    
    if not business:
        return "Business not found", 404
    
    return render_template(
        "business_dashboard.html",
        business=business,
        branches=branches,
        stats=stats,
        invoices=invoices,
        purchase_orders=purchase_orders
    )


@app.route("/business/branches")
def business_branches():
    """View and manage branches"""
    
    if not session.get('user') or not session.get('business_id'):
        return redirect(url_for('login'))
    
    from services.business_service import get_business_branches
    
    business_id = session.get('business_id')
    branches = get_business_branches(business_id)
    
    return render_template("business_branches.html", branches=branches)


@app.route("/business/inventory")
def business_inventory():
    """View and manage inventory"""
    
    if not session.get('user') or not session.get('business_id'):
        return redirect(url_for('login'))
    
    from services.business_service import get_business_branches, get_branch_inventory, get_business_products
    
    business_id = session.get('business_id')
    branch_id = request.args.get('branch_id', type=int)
    
    branches = get_business_branches(business_id)
    products = get_business_products(business_id)
    
    if branch_id:
        inventory = get_branch_inventory(branch_id)
    else:
        inventory = []
    
    return render_template(
        "business_inventory.html",
        branches=branches,
        inventory=inventory,
        products=products,
        selected_branch_id=branch_id
    )


@app.route("/business/products")
def business_products():
    """View and manage products"""
    
    if not session.get('user') or not session.get('business_id'):
        return redirect(url_for('login'))
    
    from services.business_service import get_business_products
    
    business_id = session.get('business_id')
    products = get_business_products(business_id)
    
    return render_template("business_products.html", products=products)


@app.route("/business/suppliers")
def business_suppliers():
    """View and manage suppliers"""
    
    if not session.get('user') or not session.get('business_id'):
        return redirect(url_for('login'))
    
    from services.business_service import get_business_suppliers
    
    business_id = session.get('business_id')
    suppliers = get_business_suppliers(business_id)
    
    return render_template("business_suppliers.html", suppliers=suppliers)


@app.route("/business/invoices")
def business_invoices():
    """Create and track sales invoices"""

    if not session.get('user') or not session.get('business_id'):
        return redirect(url_for('login'))

    from services.business_service import (
        get_business_branches,
        get_business_products,
        get_business_invoices,
    )

    business_id = session.get('business_id')
    return render_template(
        "business_invoices.html",
        branches=get_business_branches(business_id),
        products=get_business_products(business_id),
        invoices=get_business_invoices(business_id),
    )


@app.route("/business/purchases")
def business_purchases():
    """Create and track purchase orders"""

    if not session.get('user') or not session.get('business_id'):
        return redirect(url_for('login'))

    from services.business_service import (
        get_business_branches,
        get_business_products,
        get_business_purchase_orders,
        get_business_suppliers,
    )

    business_id = session.get('business_id')
    return render_template(
        "business_purchases.html",
        branches=get_business_branches(business_id),
        products=get_business_products(business_id),
        suppliers=get_business_suppliers(business_id),
        purchase_orders=get_business_purchase_orders(business_id),
    )


@app.route("/business/gst")
def business_gst():
    """GST input/output tax report"""

    if not session.get('user') or not session.get('business_id'):
        return redirect(url_for('login'))

    from services.business_service import get_gst_report

    business_id = session.get('business_id')
    report = get_gst_report(business_id)

    return render_template("business_gst.html", report=report)


@app.route("/business/api/create-branch", methods=["POST"])
def api_create_branch():
    """API to create a new branch"""
    
    if not session.get('user') or not session.get('business_id'):
        return jsonify(success=False, error="Unauthorized"), 401
    
    from services.business_service import create_branch
    
    data = get_json_data()
    business_id = session.get('business_id')
    
    branch_name = data.get('branch_name', '').strip()
    address = data.get('address', '').strip()
    city = data.get('city', '').strip()
    state = data.get('state', '').strip()
    pincode = data.get('pincode', '').strip()
    phone = data.get('phone', '').strip()
    
    if not branch_name:
        return jsonify(success=False, error="Branch name is required"), 400
    
    result = create_branch(
        business_id=business_id,
        branch_name=branch_name,
        address=address,
        city=city,
        state=state,
        pincode=pincode,
        phone=phone
    )
    
    return jsonify(result)


@app.route("/business/api/add-product", methods=["POST"])
def api_add_product():
    """API to add a new product"""
    
    if not session.get('user') or not session.get('business_id'):
        return jsonify(success=False, error="Unauthorized"), 401
    
    from services.business_service import add_product
    
    data = get_json_data()
    business_id = session.get('business_id')
    
    product_name = data.get('product_name', '').strip()
    sku = data.get('sku', '').strip()
    category = data.get('category', '').strip()
    unit_price = data.get('unit_price', 0)
    cost_price = data.get('cost_price', 0)
    tax_rate = data.get('tax_rate', 0)
    description = data.get('description', '').strip()
    
    if not product_name or not sku:
        return jsonify(success=False, error="Product name and SKU are required"), 400
    
    result = add_product(
        business_id=business_id,
        product_name=product_name,
        sku=sku,
        category=category,
        unit_price=float(unit_price) if unit_price else 0,
        cost_price=float(cost_price) if cost_price else 0,
        tax_rate=float(tax_rate) if tax_rate else 0,
        description=description
    )
    
    return jsonify(result)


@app.route("/business/api/add-supplier", methods=["POST"])
def api_add_supplier():
    """API to add a new supplier"""
    
    if not session.get('user') or not session.get('business_id'):
        return jsonify(success=False, error="Unauthorized"), 401
    
    from services.business_service import add_supplier
    
    data = get_json_data()
    business_id = session.get('business_id')
    
    supplier_name = data.get('supplier_name', '').strip()
    contact_person = data.get('contact_person', '').strip()
    email = data.get('email', '').strip()
    phone = data.get('phone', '').strip()
    address = data.get('address', '').strip()
    city = data.get('city', '').strip()
    state = data.get('state', '').strip()
    pincode = data.get('pincode', '').strip()
    gst_id = data.get('gst_id', '').strip()
    payment_terms = data.get('payment_terms', '').strip()
    
    if not supplier_name:
        return jsonify(success=False, error="Supplier name is required"), 400
    
    result = add_supplier(
        business_id=business_id,
        supplier_name=supplier_name,
        contact_person=contact_person,
        email=email,
        phone=phone,
        address=address,
        city=city,
        state=state,
        pincode=pincode,
        gst_id=gst_id,
        payment_terms=payment_terms
    )
    
    return jsonify(result)


@app.route("/business/api/update-stock", methods=["POST"])
def api_update_stock():
    """API to adjust stock manually"""

    if not session.get('user') or not session.get('business_id'):
        return jsonify(success=False, error="Unauthorized"), 401

    from services.business_service import update_stock

    data = get_json_data()
    branch_id = data.get('branch_id')
    product_id = data.get('product_id')
    quantity = data.get('quantity')
    reorder_level = data.get('reorder_level')
    notes = data.get('notes', '').strip()

    if not branch_id or not product_id or quantity is None:
        return jsonify(success=False, error="Branch, product, and quantity are required"), 400

    result = update_stock(
        branch_id=int(branch_id),
        product_id=int(product_id),
        new_quantity=float(quantity),
        transaction_type="adjustment",
        reference_id="manual",
        notes=notes
    )

    if result.get("success") and reorder_level not in (None, ""):
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            "UPDATE inventory SET reorder_level=? WHERE id=?",
            (float(reorder_level), result["inventory_id"]),
        )
        conn.commit()
        conn.close()

    return jsonify(result)


@app.route("/business/api/create-invoice", methods=["POST"])
def api_create_invoice():
    """API to create a sales invoice"""

    if not session.get('user') or not session.get('business_id'):
        return jsonify(success=False, error="Unauthorized"), 401

    from services.business_service import create_invoice

    data = get_json_data()
    business_id = session.get('business_id')

    customer_name = data.get('customer_name', '').strip()
    branch_id = data.get('branch_id')
    product_id = data.get('product_id')
    quantity = data.get('quantity')

    if not customer_name or not branch_id or not product_id or not quantity:
        return jsonify(success=False, error="Customer, branch, product, and quantity are required"), 400

    result = create_invoice(
        business_id=business_id,
        branch_id=int(branch_id),
        customer_name=customer_name,
        customer_email=data.get('customer_email', '').strip(),
        customer_phone=data.get('customer_phone', '').strip(),
        product_id=int(product_id),
        quantity=float(quantity),
        due_date=data.get('due_date', '').strip(),
        notes=data.get('notes', '').strip(),
    )

    return jsonify(result)


@app.route("/business/api/mark-invoice-paid", methods=["POST"])
def api_mark_invoice_paid():
    """API to mark an invoice paid"""

    if not session.get('user') or not session.get('business_id'):
        return jsonify(success=False, error="Unauthorized"), 401

    from services.business_service import mark_invoice_paid

    data = get_json_data()
    invoice_id = data.get('invoice_id')
    if not invoice_id:
        return jsonify(success=False, error="Invoice ID is required"), 400

    return jsonify(mark_invoice_paid(int(invoice_id), session.get('business_id')))


@app.route("/business/api/create-purchase", methods=["POST"])
def api_create_purchase():
    """API to create a received purchase order"""

    if not session.get('user') or not session.get('business_id'):
        return jsonify(success=False, error="Unauthorized"), 401

    from services.business_service import create_purchase_order

    data = get_json_data()
    required = ["supplier_id", "branch_id", "product_id", "quantity", "unit_price"]
    if any(not data.get(field) for field in required):
        return jsonify(success=False, error="Supplier, branch, product, quantity, and unit price are required"), 400

    result = create_purchase_order(
        business_id=session.get('business_id'),
        supplier_id=int(data.get('supplier_id')),
        branch_id=int(data.get('branch_id')),
        product_id=int(data.get('product_id')),
        quantity=float(data.get('quantity')),
        unit_price=float(data.get('unit_price')),
        expected_delivery_date=data.get('expected_delivery_date', '').strip(),
        notes=data.get('notes', '').strip(),
    )

    return jsonify(result)


@app.route("/api/register", methods=["POST"])
def api_register():

    data = request.get_json(silent=True) or {}

    username = data.get("username", "").strip()

    email = data.get("email", "").strip()

    password = data.get("password", "").strip()

    confirm_password = data.get("confirm_password", "").strip()

    role = data.get("role", "household")

    group_name = (
        data.get("group_name", "").strip()
        or username
    )

    if not username or not email or not password or not confirm_password:

        return jsonify(
            success=False,
            error="Username, Gmail, password, and confirm password are required"
        ), 400

    if not valid_gmail(email):

        return jsonify(
            success=False,
            error="Please provide a valid Gmail address"
        ), 400

    if not valid_username(username):

        return jsonify(
            success=False,
            error="Username must contain at least 3 characters"
        ), 400

    if password != confirm_password:

        return jsonify(
            success=False,
            error="Password and confirm password must match"
        ), 400

    if not valid_password(password):

        return jsonify(
            success=False,
            error="""
Password must contain:
1 Capital Letter
1 Small Letter
1 Number
1 Special Character
Minimum 8 Characters
"""
        ), 400

    # HASH USER PASSWORD
    hashed_password = generate_password_hash(password)

    conn = get_db()

    cur = conn.cursor()

    cur.execute(
        "SELECT id FROM users WHERE username=? OR email=?",
        (username, email)
    )

    existing_user = cur.fetchone()

    if existing_user:

        conn.close()

        return jsonify(
            success=False,
            error="Username or Gmail already exists"
        ), 409

    registration_date = date.today().isoformat()

    cur.execute(
        """
        INSERT INTO users(

            username,
            email,
            password,
            role,
            group_name,
            registration_date

        )
        VALUES(?,?,?,?,?,?)
        """,
        (
            username,
            email,
            hashed_password,
            role,
            group_name,
            registration_date
        ),
    )

    conn.commit()

    conn.close()

    return jsonify(
        success=True,
        username=username,
        email=email,
        role=role,
        group_name=group_name
    )


# =========================================================
# API LOGIN
# =========================================================

@app.route("/api/login", methods=["POST"])
def api_login():

    data = request.get_json(silent=True) or {}

    username = data.get("username", "").strip()

    password = data.get("password", "").strip()

    if not username or not password:

        return jsonify(
            success=False,
            error="Username and password are required"
        ), 400

    conn = get_db()

    cur = conn.cursor()

    cur.execute(
        "SELECT * FROM users WHERE username=? OR email=?",
        (username, username)
    )

    user = cur.fetchone()

    conn.close()

    # CHECK HASHED PASSWORD
    if not user or not check_password_hash(
        user["password"],
        password
    ):

        return jsonify(
            success=False,
            error="Invalid credentials"
        ), 401

    session["user"] = user["username"]

    session["role"] = user["role"]

    session["group_name"] = user["group_name"]

    return jsonify(
        success=True,
        username=user["username"],
        role=user["role"],
        group_name=user["group_name"],
    )


@app.route('/api/auth/token', methods=['POST'])
def api_auth_token():
    data = request.get_json(silent=True) or {}
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()

    if not username or not password:
        return jsonify(success=False, error='username/password required'), 400

    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT * FROM users WHERE username=? OR email=?', (username, username))
    user = cur.fetchone()
    conn.close()

    if not user or not check_password_hash(user['password'], password):
        return jsonify(success=False, error='Invalid credentials'), 401

    token = create_jwt(user['username'], user['group_name'], user['role'])

    return jsonify(success=True, token=token, username=user['username'], role=user['role'], group_name=user['group_name'])



# ==========================
# AI INSIGHTS API
# ==========================

@app.route('/api/ai/insights', methods=['POST'])
def api_ai_insights():

    # allow JWT or session
    auth_ok = False
    auth = request.headers.get('Authorization','')
    if auth.startswith('Bearer '):
        payload = decode_jwt(auth.split(' ',1)[1].strip())
        if payload:
            auth_ok = True
            group_name = payload.get('group')
    if not auth_ok and 'user' in session:
        auth_ok = True
        group_name = session['group_name']

    if not auth_ok:
        return jsonify(success=False, error='Authentication required'), 401

    data = request.get_json(silent=True) or {}
    date_from = data.get('date_from')
    date_to = data.get('date_to')

    # group_name obtained above

    conn = get_db()
    cur = conn.cursor()

    query = "SELECT date, user, type, amount, category, note FROM transactions WHERE group_name=?"
    params = [group_name]
    if date_from:
        query += ' AND date >= ?'
        params.append(date_from)
    if date_to:
        query += ' AND date <= ?'
        params.append(date_to)

    cur.execute(query + ' ORDER BY id DESC', tuple(params))
    rows = cur.fetchall()
    conn.close()

    txs = []
    for r in rows:
        txs.append({
            'date': r['date'],
            'user': r['user'],
            'type': r['type'],
            'amount': r['amount'],
            'category': r['category'],
            'note': r['note']
        })

    insights = generate_insights(txs)

    return jsonify(success=True, insights=insights)


# ==========================
# OCR UPLOAD API
# ==========================

@app.route('/api/ocr/scan', methods=['POST'])
def api_ocr_scan():

    # allow JWT or session
    auth_ok = False
    auth = request.headers.get('Authorization','')
    if auth.startswith('Bearer '):
        payload = decode_jwt(auth.split(' ',1)[1].strip())
        if payload:
            auth_ok = True
    if not auth_ok and 'user' in session:
        auth_ok = True

    if not auth_ok:
        return jsonify(success=False, error='Authentication required'), 401

    if 'file' not in request.files:
        return jsonify(success=False, error='File is required'), 400

    f = request.files['file']
    content = f.read()

    from services.ocr_service import ocr_image

    result = ocr_image(content)

    return jsonify(success=True, result=result)


# =========================================================
# DASHBOARD
# =========================================================

def render_simple_page(title: str, content: str) -> str:
    template = '''
    <!DOCTYPE html>
    <html>
    <head>
      <meta charset="utf-8">
      <meta name="viewport" content="width=device-width, initial-scale=1.0">
      <title>{{ title }}</title>
      <link rel="stylesheet" href="/static/style.css">
    </head>
    <body class="dashboard-page">
      <aside class="sidebar">
        <div class="sidebar-logo">
          <div class="logo-icon">💰</div>
          <span class="logo-text">ExpensesIQ</span>
        </div>
        <nav class="sidebar-nav">
          <a href="{{ url_for('dashboard') }}" class="nav-item"> <span class="nav-icon">📊</span> <span class="nav-text">Dashboard</span> </a>
          <a href="{{ url_for('transactions_page') }}" class="nav-item"> <span class="nav-icon">💳</span> <span class="nav-text">Transactions</span> </a>
          <a href="{{ url_for('budgets') }}" class="nav-item"> <span class="nav-icon">🎯</span> <span class="nav-text">Budgets</span> </a>
          <a href="{{ url_for('goals_page') }}" class="nav-item"> <span class="nav-icon">🎪</span> <span class="nav-text">Goals</span> </a>
          <a href="{{ url_for('calendar_page') }}" class="nav-item"> <span class="nav-icon">📅</span> <span class="nav-text">Calendar</span> </a>
        </nav>
        <div class="sidebar-footer">
          <a href="/logout" class="logout-btn">
            <span class="nav-icon">🚪</span>
            <span class="nav-text">Logout</span>
          </a>
        </div>
      </aside>
      <main class="dashboard-main">
        <header class="dashboard-header">
          <div class="header-left">
            <h1>{{ title }}</h1>
            <span class="header-subtitle">Welcome back, {{ user }}!</span>
          </div>
        </header>
        <section class="dashboard-content">
          <div class="dashboard-card">
            <div class="card-header">
              <h2>{{ title }}</h2>
            </div>
            <div class="content-body">{{ content|safe }}</div>
          </div>
        </section>
      </main>
    </body>
    </html>
    '''
    return render_template_string(template, title=title, content=content, user=session.get('user', 'Guest'))

@app.route("/dashboard", methods=["GET", "POST"])
def dashboard():

    if "user" not in session:
        return redirect(url_for("login"))

    user = session["user"]
    role = session.get("role", "household")
    group_name = session.get("group_name")
    member_name = session.get("member_name", user)
    day_limit = session.get("daily_limit", 1000)

    conn = get_db()
    cur = conn.cursor()

    if request.method == "POST":
        action = request.form.get("action", "transaction")

        if action == "transaction":
            amount = request.form.get("amount")
            category = request.form.get("category", "").strip()
            if category == "Other":
                category = request.form.get("category_other", "").strip() or category
            transaction_type = request.form.get("type", "expense")
            note = request.form.get("note", "").strip()
            transaction_date = request.form.get("transaction_date", "").strip()

            if not valid_amount(amount):
                conn.close()
                return redirect(url_for("dashboard"))

            txn_date = date.today().isoformat()
            if transaction_date:
                try:
                    datetime.fromisoformat(transaction_date)
                    txn_date = transaction_date
                except ValueError:
                    pass

            cur.execute(
                "INSERT INTO transactions (group_name, user, amount, category, type, note, date) VALUES (?,?,?,?,?,?,?)",
                (
                    group_name,
                    member_name,
                    float(amount),
                    category,
                    transaction_type,
                    note,
                    txn_date,
                ),
            )
            conn.commit()
            conn.close()
            return redirect(url_for("dashboard"))

        if action == "goal":
            title = request.form.get("title", "").strip()
            target = request.form.get("target")
            saved = request.form.get("saved", "0")
            due_date = request.form.get("due_date")

            try:
                target_value = float(target)
                saved_value = float(saved)
            except Exception:
                conn.close()
                return redirect(url_for("dashboard"))

            if not title or target_value <= 0 or saved_value < 0:
                conn.close()
                return redirect(url_for("dashboard"))

            cur.execute(
                "INSERT INTO goals (group_name, user, title, target, saved, due_date) VALUES (?,?,?,?,?,?)",
                (
                    group_name,
                    member_name,
                    title,
                    target_value,
                    saved_value,
                    due_date,
                ),
            )
            conn.commit()
            conn.close()
            return redirect(url_for("dashboard"))

        if action == "goal_update":
            goal_id = request.form.get("goal_id")
            target = request.form.get("target")
            saved = request.form.get("saved", "0")
            due_date = request.form.get("due_date")

            try:
                goal_id_value = int(goal_id)
                target_value = float(target)
                saved_value = float(saved)
            except Exception:
                conn.close()
                return redirect(url_for("dashboard"))

            if target_value <= 0 or saved_value < 0:
                conn.close()
                return redirect(url_for("dashboard"))

            cur.execute(
                """
                UPDATE goals
                SET target=?, saved=?, due_date=?
                WHERE id=? AND group_name=?
                """,
                (
                    target_value,
                    saved_value,
                    due_date,
                    goal_id_value,
                    group_name,
                ),
            )
            conn.commit()
            conn.close()
            return redirect(url_for("dashboard"))

        if action == "set_limit":
            daily_limit = request.form.get("daily_limit")
            if valid_amount(daily_limit):
                session["daily_limit"] = float(daily_limit)
                cur.execute(
                    "UPDATE users SET daily_limit=? WHERE username=?",
                    (float(daily_limit), user),
                )
                conn.commit()
            conn.close()
            return redirect(url_for("dashboard"))

        if action == "payment":
            amount = request.form.get("pay_amount")
            direction = request.form.get("direction", "out")
            description = request.form.get("description", "").strip()
            status = request.form.get("status", "due")
            due_date_payment = request.form.get("due_date_payment")

            if not valid_amount(amount) or not description:
                conn.close()
                return redirect(url_for("dashboard"))

            cur.execute(
                "INSERT INTO payments (group_name, user, amount, direction, description, status, due_date) VALUES (?,?,?,?,?,?,?)",
                (
                    group_name,
                    member_name,
                    float(amount),
                    direction,
                    description,
                    status,
                    due_date_payment,
                ),
            )
            conn.commit()
            conn.close()
            return redirect(url_for("dashboard"))

    # PAGINATION
    try:
        page = int(request.args.get('page', 1))
        if page < 1:
            page = 1
    except ValueError:
        page = 1

    per_page = 10
    offset = (page - 1) * per_page

    filter_from = request.args.get('date_from', '')
    filter_to = request.args.get('date_to', '')

    query = "SELECT * FROM transactions WHERE group_name=?"
    params = [group_name]
    if filter_from:
        query += " AND date >= ?"
        params.append(filter_from)
    if filter_to:
        query += " AND date <= ?"
        params.append(filter_to)

    query += " ORDER BY id DESC"
    cur.execute(query, tuple(params))
    all_transactions = cur.fetchall()

    total_count = len(all_transactions)
    total_pages = max(1, (total_count + per_page - 1) // per_page)
    transactions = all_transactions[offset:offset + per_page]

    total_income = sum(
        t["amount"]
        for t in all_transactions
        if t["type"] == "income"
    )
    total_expense = sum(
        t["amount"]
        for t in all_transactions
        if t["type"] == "expense"
    )
    net_balance = total_income - total_expense

    today = date.today().isoformat()
    today_expense = sum(
        t["amount"]
        for t in all_transactions
        if t["type"] == "expense" and t["date"] == today
    )

    week_expense = sum(
        t["amount"]
        for t in all_transactions
        if t["type"] == "expense"
        and datetime.fromisoformat(t["date"]).date() >= (date.today() - timedelta(days=6))
    )

    month_expense = sum(
        t["amount"]
        for t in all_transactions
        if t["type"] == "expense"
        and datetime.fromisoformat(t["date"]).date().month == date.today().month
    )

    # Determine monthly budget default from used monthly expense
    # Use session `monthly_budget` if set, otherwise fall back to actual used `month_expense`.
    monthly_budget = session.get('monthly_budget', month_expense)
    # If user hasn't set a daily limit, derive it from the monthly budget
    # by dividing evenly across the days in the current month.
    if 'daily_limit' not in session:
        days_in_month = calendar.monthrange(date.today().year, date.today().month)[1]
        # avoid division by zero
        day_limit = (monthly_budget / max(1, days_in_month)) if monthly_budget else 0

    # Aggregate totals per (category, type) so income and expense categories are both shown
    category_totals = {}
    for t in all_transactions:
        category = (t["category"] or "Uncategorized")
        typ = t["type"] if t["type"] else "expense"
        key = (category, typ)
        category_totals[key] = category_totals.get(key, 0) + t["amount"]

    category_data = [
        {"category": cat, "amount": round(amount, 2), "type": typ}
        for (cat, typ), amount in category_totals.items()
    ]

    # Prepare monthly trends data (income and expenses) for the selected range
    # Determine start and end dates for trends
    try:
        if filter_from:
            trends_start = datetime.fromisoformat(filter_from).date()
        else:
            # default to 5 months ago
            today_dt = date.today()
            month = today_dt.month - 5
            year = today_dt.year
            while month <= 0:
                month += 12
                year -= 1
            trends_start = date(year, month, 1)

        if filter_to:
            trends_end = datetime.fromisoformat(filter_to).date()
        else:
            trends_end = date.today()
    except Exception:
        # on parse error fallback to last 6 months
        trends_end = date.today()
        trends_start = (trends_end - timedelta(days=30 * 5))

    # build month list
    months = []
    cur_month = date(trends_start.year, trends_start.month, 1)
    while cur_month <= trends_end:
        months.append((cur_month.year, cur_month.month))
        # increment month
        y = cur_month.year + (cur_month.month // 12)
        m = cur_month.month % 12 + 1
        cur_month = date(y, m, 1)

    trends_labels = []
    trends_income = []
    trends_expenses = []

    for y, m in months:
        # month label
        trends_labels.append(f"{calendar.month_abbr[m]} {y}")
        # sum amounts for this month
        inc = 0.0
        exp = 0.0
        for t in all_transactions:
            try:
                td = datetime.fromisoformat(t['date']).date()
            except Exception:
                continue
            if td.year == y and td.month == m:
                if t['type'] == 'income':
                    inc += t['amount']
                elif t['type'] == 'expense':
                    exp += t['amount']
        trends_income.append(round(inc, 2))
        trends_expenses.append(round(exp, 2))

    # Build chart data for pie chart (Expense Overview)
    expenses_only = [d for d in category_data if d['type'] == 'expense']
    chart_labels = [d['category'] for d in expenses_only]
    chart_amounts = [d['amount'] for d in expenses_only]
    
    chart_data = {
        "labels": chart_labels,
        "datasets": [{
            "data": chart_amounts,
            "backgroundColor": [
                "#FF6B6B", "#4ECDC4", "#45B7D1", "#FFA07A", "#98D8C8",
                "#F7DC6F", "#BB8FCE", "#85C1E2", "#F8B88B", "#A8E6CF"
            ],
            "borderColor": "var(--surface)",
            "borderWidth": 2
        }]
    }
    
    # Build chart data for line chart (Monthly Trends)
    trends_data = {
        "labels": trends_labels,
        "datasets": [
            {
                "label": "Income",
                "data": trends_income,
                "borderColor": "#22C55E",
                "backgroundColor": "rgba(34, 197, 94, 0.1)",
                "borderWidth": 2,
                "fill": False,
                "tension": 0.4,
                "pointRadius": 4,
                "pointBackgroundColor": "#22C55E"
            },
            {
                "label": "Expenses",
                "data": trends_expenses,
                "borderColor": "#EF4444",
                "backgroundColor": "rgba(239, 68, 68, 0.1)",
                "borderWidth": 2,
                "fill": False,
                "tension": 0.4,
                "pointRadius": 4,
                "pointBackgroundColor": "#EF4444"
            }
        ]
    }

    alert = ""
    if role != "business" and today_expense > float(day_limit):
        alert = "Warning: Daily expense limit crossed!"

    goals = []
    cur.execute(
        "SELECT * FROM goals WHERE group_name=? ORDER BY id DESC",
        (group_name,)
    )
    for goal in cur.fetchall():
        target = goal["target"] or 0
        saved = goal["saved"] or 0
        progress = int(min(100, (saved / target * 100) if target else 0))
        goals.append({
            "id": goal["id"],
            "title": goal["title"],
            "target": target,
            "saved": saved,
            "due_date": goal["due_date"],
            "progress": progress,
        })

    payments = []
    show_business_features = (role == "business")
    if show_business_features:
        cur.execute(
            "SELECT * FROM payments WHERE group_name=? ORDER BY id DESC",
            (group_name,)
        )
        payments = [dict(p) for p in cur.fetchall()]

    conn.close()

    return render_template(
        "dashboard_redesigned.html",
        transactions=transactions,
        page=page,
        total_pages=total_pages,
        total_income=round(total_income, 2),
        total_expense=round(total_expense, 2),
        net_balance=round(net_balance, 2),
        balance=round(net_balance, 2),
        today_expense=round(today_expense, 2),
        week_expense=round(week_expense, 2),
        month_expense=round(month_expense, 2),
        category_data=category_data,
        chart_data=json.dumps(chart_data),
        trends_data=json.dumps(trends_data),
        trends_labels=trends_labels,
        trends_income=trends_income,
        trends_expenses=trends_expenses,
        alert=alert,
        user=user,
        role=role,
        group_name=group_name,
        member_name=member_name,
        day_limit=round(float(day_limit), 2),
        member_limit=round(float(day_limit), 2),
        show_business_features=show_business_features,
        goals=goals,
        payments=payments,
        filter_from=filter_from,
        filter_to=filter_to,
        monthly_budget=round(float(monthly_budget), 2),
        today_date=date.today().isoformat(),
    )

@app.route('/test-otp')
def test_otp():
    """Test page for OTP SMS functionality"""
    return render_template('test_otp.html')

@app.route('/transactions')
def transactions_page():
    if 'user' not in session:
        return redirect(url_for('login'))

    group_name = session['group_name']
    requested_date = request.args.get('date', '').strip()
    conn = get_db()
    cur = conn.cursor()

    if requested_date:
        cur.execute(
            'SELECT * FROM transactions WHERE group_name=? AND date=? ORDER BY id DESC',
            (group_name, requested_date),
        )
        rows = [dict(r) for r in cur.fetchall()]
        total_for_date = sum(r.get('amount', 0.0) for r in rows if r.get('type') == 'expense')
    else:
        cur.execute('SELECT * FROM transactions WHERE group_name=? ORDER BY id DESC', (group_name,))
        rows = [dict(r) for r in cur.fetchall()]
        total_for_date = None

    conn.close()

    return render_template(
        'transactions.html',
        transactions=rows,
        user=session.get('user'),
        date_filter=requested_date,
        total_for_date=total_for_date,
    )

@app.route('/goals', methods=['GET', 'POST'])
def goals_page():
    if 'user' not in session:
        return redirect(url_for('login'))

    group_name = session['group_name']
    conn = get_db()
    cur = conn.cursor()

    if request.method == 'POST':
        goal_id = request.form.get('goal_id')
        target = request.form.get('target')
        saved = request.form.get('saved', '0')
        due_date = request.form.get('due_date')

        try:
            goal_id_value = int(goal_id)
            target_value = float(target)
            saved_value = float(saved)
        except Exception:
            conn.close()
            return redirect(url_for('goals_page'))

        if target_value > 0 and saved_value >= 0:
            cur.execute(
                'UPDATE goals SET target=?, saved=?, due_date=? WHERE id=? AND group_name=?',
                (target_value, saved_value, due_date, goal_id_value, group_name),
            )
            conn.commit()

        conn.close()
        return redirect(url_for('goals_page'))

    cur.execute('SELECT * FROM goals WHERE group_name=? ORDER BY id DESC', (group_name,))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()

    goals = []
    for g in rows:
        progress = int(min(100, (g['saved'] / g['target'] * 100) if g['target'] else 0))
        goals.append({
            'id': g['id'],
            'title': g['title'],
            'target': g['target'],
            'saved': g['saved'],
            'due_date': g['due_date'],
            'progress': progress,
        })

    return render_template('goals.html', goals=goals)

@app.route('/budgets', methods=['GET', 'POST'])
def budgets():
    if 'user' not in session:
        return redirect(url_for('login'))

    user = session['user']
    group_name = session.get('group_name')
    conn = get_db()
    cur = conn.cursor()

    # compute used monthly expense
    cur.execute('SELECT * FROM transactions WHERE group_name=?', (group_name,))
    txs = [dict(t) for t in cur.fetchall()]
    month_expense = sum(
        t['amount'] for t in txs
        if t.get('type') == 'expense' and datetime.fromisoformat(t.get('date')).date().month == date.today().month
    )

    # load current stored budget for user
    cur.execute('SELECT monthly_budget FROM users WHERE username=?', (user,))
    row = cur.fetchone()
    stored_budget = row['monthly_budget'] if row is not None else None

    if request.method == 'POST':
        # set/update monthly budget for this user
        mb = request.form.get('monthly_budget', '').strip()
        try:
            mb_val = float(mb) if mb else None
        except Exception:
            mb_val = None

        cur.execute('UPDATE users SET monthly_budget=? WHERE username=?', (mb_val, user))
        conn.commit()
        session['monthly_budget'] = mb_val
        stored_budget = mb_val

        return redirect(url_for('budgets'))

    # derive daily limit from stored monthly budget or used month expense
    monthly_budget = stored_budget if stored_budget is not None else month_expense
    days_in_month = calendar.monthrange(date.today().year, date.today().month)[1]
    derived_daily = (monthly_budget / max(1, days_in_month)) if monthly_budget else 0

    if stored_budget is not None and monthly_budget > 0:
        progress_pct = round(min(100, (month_expense / stored_budget) * 100), 1)
        remaining_amount = stored_budget - month_expense
        budget_status = 'On track' if month_expense <= stored_budget else 'Over budget'
        status_text = 'left' if remaining_amount >= 0 else 'over'
        remaining_label = f"₹{abs(round(remaining_amount,2)):.2f} {status_text}"
    else:
        progress_pct = 100
        remaining_amount = 0
        budget_status = 'Budget not set'
        remaining_label = 'Set a target to track progress'

    conn.close()

    budget_display = f'₹{round(float(stored_budget),2):,.2f}' if stored_budget is not None else 'Not set'

    html = f"""
    <div class="budget-hero">
      <div>
        <h3>Smart budget planning</h3>
        <p>Stay ahead of your spending with a clear monthly snapshot, daily guidance, and a progress meter designed for easy review.</p>
      </div>
      <div class="budget-summary-pill">
        <span>Monthly target</span>
        <strong>{budget_display}</strong>
      </div>
    </div>

    <div class="dashboard-grid budget-metrics">
      <div class="budget-stat">
        <span>Used this month</span>
        <div class="stat-value">₹{round(month_expense,2):,.2f}</div>
        <div class="stat-note">Current expense total for {date.today():%B %Y}</div>
      </div>
      <div class="budget-stat">
        <span>Monthly budget</span>
        <div class="stat-value">{budget_display}</div>
        <div class="stat-note">Your target amount</div>
      </div>
      <div class="budget-stat">
        <span>Daily guidance</span>
        <div class="stat-value">₹{round(float(derived_daily),2):,.2f}</div>
        <div class="stat-note">Suggested daily spend based on budget</div>
      </div>
    </div>

    <div class="budget-progress">
      <div class="progress-header">
        <div>
          <div class="progress-title">Budget progress</div>
          <div class="progress-label">{budget_status}</div>
        </div>
        <div><strong>{progress_pct}%</strong></div>
      </div>
      <div class="budget-progress-bar">
        <div class="budget-progress-fill" style="width:{progress_pct}%;"></div>
      </div>
      <div class="budget-progress-info">
        <p><strong>{budget_display}</strong> target</p>
        <p><strong>{remaining_label}</strong> remaining</p>
      </div>
    </div>

    <div class="dashboard-card budget-form-card">
      <h3>Update monthly budget</h3>
      <form method="post" class="dashboard-form">
        <div class="form-group">
          <label>Monthly budget (₹)
            <input class="form-input" name="monthly_budget" type="number" step="0.01" min="0" value="{stored_budget if stored_budget is not None else ''}">
          </label>
        </div>
        <button class="form-btn primary" type="submit">Save budget</button>
      </form>
      <div class="budget-tip">Leave this blank to continue using actual spending as your monthly benchmark.</div>
    </div>
    """

    return render_simple_page('Budgets', html)

@app.route('/calendar')
def calendar_page():
    if 'user' not in session:
        return redirect(url_for('login'))

    # show a simple monthly calendar with per-day expense totals
    user = session['user']
    group_name = session.get('group_name')

    # determine month/year to display from query params
    try:
        year = int(request.args.get('year', date.today().year))
        month = int(request.args.get('month', date.today().month))
    except Exception:
        year = date.today().year
        month = date.today().month

    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT * FROM transactions WHERE group_name=?', (group_name,))
    txs = [dict(t) for t in cur.fetchall()]

    # prepare per-day totals and transaction counts for the month
    days_in_month = calendar.monthrange(year, month)[1]
    day_info = {d: {'total': 0.0, 'count': 0} for d in range(1, days_in_month + 1)}
    for t in txs:
        if t.get('type') != 'expense':
            continue
        try:
            td = datetime.fromisoformat(t.get('date')).date()
        except Exception:
            continue
        if td.year == year and td.month == month:
            info = day_info.get(td.day)
            if info is not None:
                info['total'] += t.get('amount', 0.0)
                info['count'] += 1

    # navigation months
    prev_month = month - 1
    prev_year = year
    if prev_month < 1:
        prev_month = 12
        prev_year -= 1

    next_month = month + 1
    next_year = year
    if next_month > 12:
        next_month = 1
        next_year += 1

    conn.close()

    return render_template(
        'calendar.html',
        year=year,
        month=month,
        month_name=calendar.month_name[month],
        days_in_month=days_in_month,
        day_info=day_info,
        prev_year=prev_year,
        prev_month=prev_month,
        next_year=next_year,
        next_month=next_month,
        user=user,
        calendar=calendar,
        date=date,
    )

@app.route('/profile')
def profile():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    user = session["user"]
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT username, email FROM users WHERE username=?", (user,))
    user_data = cur.fetchone()
    conn.close()
    
    email = user_data["email"] if user_data else ""
    
    return render_template(
        "profile.html",
        user=user,
        email=email,
        first_name="",
        last_name="",
        phone="",
        member_since="2024"
    )

@app.route('/settings')
def settings():
    if 'user' not in session:
        return redirect(url_for('login'))
    return render_simple_page('Settings', '<p>Application settings will appear here.</p>')


@app.route('/preferences')
def preferences():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    return render_template("preferences.html")


@app.route('/help-support')
def help_support():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    return render_template("help_support.html")


@app.route('/set_member', methods=['POST'])
def set_member():
    if 'user' not in session:
        return redirect(url_for('login'))

    member_name = request.form.get('member_name', '').strip()
    if member_name:
        session['member_name'] = member_name
    else:
        session.pop('member_name', None)

    return redirect(url_for('dashboard'))


# ==========================
# EXPORT CSV
# ==========================

@app.route('/export/csv')
def export_csv():

    if 'user' not in session:
        return redirect(url_for('login'))

    group_name = session['group_name']

    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        "SELECT date, user, type, amount, category, note FROM transactions WHERE group_name=? ORDER BY id DESC",
        (group_name,)
    )

    rows = cur.fetchall()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['date', 'user', 'type', 'amount', 'category', 'note'])

    for r in rows:
        writer.writerow([r['date'], r['user'], r['type'], r['amount'], r['category'], r['note']])

    resp = Response(output.getvalue(), mimetype='text/csv')
    resp.headers['Content-Disposition'] = 'attachment; filename=transactions.csv'

    conn.close()

    return resp


# ==========================
# EXPORT PDF
# ==========================

@app.route('/export/pdf')
def export_pdf():

    if 'user' not in session:
        return redirect(url_for('login'))

    group_name = session['group_name']

    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        "SELECT date, user, type, amount, category, note FROM transactions WHERE group_name=? ORDER BY id DESC",
        (group_name,)
    )

    rows = cur.fetchall()

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    y = height - 40
    p.setFont('Helvetica-Bold', 14)
    p.drawString(40, y, f"Transactions - {group_name}")
    y -= 30
    p.setFont('Helvetica', 10)

    headers = ['Date', 'User', 'Type', 'Amount', 'Category', 'Note']
    p.drawString(40, y, ' | '.join(headers))
    y -= 18

    for r in rows:
        line = f"{r['date']} | {r['user']} | {r['type']} | {r['amount']} | {r['category']} | {r['note']}"
        p.drawString(40, y, (line[:120]))
        y -= 14
        if y < 60:
            p.showPage()
            y = height - 40
            p.setFont('Helvetica', 10)

    p.save()

    buffer.seek(0)
    resp = Response(buffer.read(), mimetype='application/pdf')
    resp.headers['Content-Disposition'] = 'attachment; filename=transactions.pdf'

    conn.close()

    return resp


# ==========================
# EXPORT EXCEL
# ==========================

@app.route('/export/xlsx')
def export_xlsx():

    if 'user' not in session:
        return redirect(url_for('login'))

    group_name = session['group_name']

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        "SELECT date, user, type, amount, category, note FROM transactions WHERE group_name=? ORDER BY id DESC",
        (group_name,)
    )

    rows = cur.fetchall()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Add headers
    headers = ['Date', 'User', 'Type', 'Amount', 'Category', 'Note']
    ws.append(headers)

    # Style header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add data rows
    for r in rows:
        ws.append([r['date'], r['user'], r['type'], r['amount'], r['category'], r['note']])

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Format amount column as currency
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = '$#,##0.00'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    resp = Response(buffer.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp.headers['Content-Disposition'] = 'attachment; filename=transactions.xlsx'

    conn.close()

    return resp


# ==========================
# REST API - Transactions
# ==========================


@app.route('/api/transactions', methods=['GET'])
@jwt_required
def api_transactions_list():
    # pagination
    try:
        page = int(request.args.get('page', 1))
    except Exception:
        page = 1
    per_page = int(request.args.get('per_page', 20))
    offset = (page - 1) * per_page

    # determine group
    if hasattr(request, 'jwt_payload'):
        group_name = request.jwt_payload.get('group')
    else:
        group_name = session.get('group_name')

    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT COUNT(*) FROM transactions WHERE group_name=?', (group_name,))
    total = cur.fetchone()[0]
    cur.execute('SELECT * FROM transactions WHERE group_name=? ORDER BY id DESC LIMIT ? OFFSET ?', (group_name, per_page, offset))
    rows = cur.fetchall()
    conn.close()

    txs = [dict(r) for r in rows]
    return jsonify(success=True, transactions=txs, total=total, page=page, per_page=per_page)


@app.route('/api/transactions', methods=['POST'])
@jwt_required
def api_transactions_create():
    data = request.get_json(silent=True) or {}
    amount = data.get('amount')
    category = data.get('category','')
    ttype = data.get('type','expense')
    note = data.get('note','')

    if amount is None:
        return jsonify(success=False, error='amount required'), 400
    try:
        amount = float(amount)
    except Exception:
        return jsonify(success=False, error='invalid amount'), 400

    group_name = request.jwt_payload.get('group') if hasattr(request, 'jwt_payload') else session.get('group_name')
    user = request.jwt_payload.get('sub') if hasattr(request, 'jwt_payload') else session.get('user')

    conn = get_db()
    cur = conn.cursor()
    cur.execute('INSERT INTO transactions (group_name, user, amount, category, type, note, date) VALUES (?,?,?,?,?,?,?)', (
        group_name, user, amount, category, ttype, note, date.today().isoformat()
    ))
    conn.commit()
    conn.close()

    return jsonify(success=True)


@app.route('/api/transactions/<int:trx_id>', methods=['PUT'])
@jwt_required
def api_transactions_update(trx_id):
    data = request.get_json(silent=True) or {}
    amount = data.get('amount')
    category = data.get('category')
    ttype = data.get('type')
    note = data.get('note')

    fields = []
    params = []
    if amount is not None:
        try:
            amount = float(amount)
            fields.append('amount=?')
            params.append(amount)
        except Exception:
            return jsonify(success=False, error='invalid amount'), 400
    if category is not None:
        fields.append('category=?')
        params.append(category)
    if ttype is not None:
        fields.append('type=?')
        params.append(ttype)
    if note is not None:
        fields.append('note=?')
        params.append(note)

    if not fields:
        return jsonify(success=False, error='nothing to update'), 400

    params.append(trx_id)

    conn = get_db()
    cur = conn.cursor()
    cur.execute(f"UPDATE transactions SET {', '.join(fields)} WHERE id=?", tuple(params))
    conn.commit()
    conn.close()

    return jsonify(success=True)


@app.route('/api/transactions/<int:trx_id>', methods=['DELETE'])
@jwt_required
def api_transactions_delete(trx_id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute('DELETE FROM transactions WHERE id=?', (trx_id,))
    conn.commit()
    conn.close()
    return jsonify(success=True)


# ==========================
# Recurring payments
# ==========================

@app.route('/api/recurring', methods=['POST'])
@jwt_required
def api_recurring_create():
    data = request.get_json(silent=True) or {}
    amount = data.get('amount')
    category = data.get('category','')
    ttype = data.get('type','expense')
    note = data.get('note','')
    start_date = data.get('start_date') or date.today().isoformat()
    frequency_days = int(data.get('frequency_days', 30))

    if amount is None:
        return jsonify(success=False, error='amount required'), 400
    try:
        amount = float(amount)
    except Exception:
        return jsonify(success=False, error='invalid amount'), 400

    group_name = request.jwt_payload.get('group') if hasattr(request, 'jwt_payload') else session.get('group_name')
    user = request.jwt_payload.get('sub') if hasattr(request, 'jwt_payload') else session.get('user')

    conn = get_db()
    cur = conn.cursor()
    cur.execute('INSERT INTO recurring_payments (group_name, user, amount, category, type, note, start_date, frequency_days, next_date, active) VALUES (?,?,?,?,?,?,?,?,?,?)', (
        group_name, user, amount, category, ttype, note, start_date, frequency_days, start_date, 1
    ))
    conn.commit()
    conn.close()

    return jsonify(success=True)


@app.route('/api/recurring', methods=['GET'])
@jwt_required
def api_recurring_list():
    group_name = request.jwt_payload.get('group') if hasattr(request, 'jwt_payload') else session.get('group_name')
    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT * FROM recurring_payments WHERE group_name=?', (group_name,))
    rows = cur.fetchall()
    conn.close()
    return jsonify(success=True, recurring=[dict(r) for r in rows])


@app.route('/api/recurring/process', methods=['POST'])
def api_recurring_process():
    # lightweight processor: create due transactions and advance next_date
    conn = get_db()
    cur = conn.cursor()
    today = date.today().isoformat()
    cur.execute('SELECT * FROM recurring_payments WHERE active=1 AND next_date<=?', (today,))
    due = cur.fetchall()

    processed = 0
    for r in due:
        # insert transaction
        cur.execute('INSERT INTO transactions (group_name, user, amount, category, type, note, date) VALUES (?,?,?,?,?,?,?)', (
            r['group_name'], r['user'], r['amount'], r['category'], r['type'], r['note'], today
        ))
        # compute next_date
        try:
            nd = datetime.fromisoformat(r['next_date']).date() + timedelta(days=r['frequency_days'])
            cur.execute('UPDATE recurring_payments SET next_date=? WHERE id=?', (nd.isoformat(), r['id']))
        except Exception:
            pass
        processed += 1

    conn.commit()
    conn.close()
    return jsonify(success=True, processed=processed)


# ==========================
# Email notifications
# ==========================

@app.route('/api/notify/send', methods=['POST'])
@jwt_required
def api_notify_send():
    data = request.get_json(silent=True) or {}
    to = data.get('to')
    subject = data.get('subject', 'Notification')
    body = data.get('body', '')
    if not to:
        return jsonify(success=False, error='to is required'), 400
    res = send_email(to, subject, body)
    return jsonify(success=True, result=res)


# ==========================
# Forecasting
# ==========================

@app.route('/api/forecast', methods=['GET'])
@jwt_required
def api_forecast():
    # collect transactions for group
    if hasattr(request, 'jwt_payload'):
        group_name = request.jwt_payload.get('group')
    else:
        group_name = session.get('group_name')

    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT date, type, amount FROM transactions WHERE group_name=?', (group_name,))
    rows = cur.fetchall()
    conn.close()

    txs = [dict(r) for r in rows]
    res = forecast_monthly_expense(txs, months=3)
    return jsonify(success=True, forecast=res)


# =========================================================
# ADD TRANSACTION
# =========================================================

@app.route("/add_transaction", methods=["POST"])
def add_transaction():

    if "user" not in session:

        return redirect(url_for("login"))

    amount = request.form["amount"]

    if not valid_amount(amount):

        return """
        <h2>Invalid Amount</h2>

        <a href='/dashboard'>Go Back</a>
        """

    amount = float(amount)

    category = request.form["category"]

    transaction_type = request.form["type"]

    note = request.form.get("note", "").strip()

    conn = get_db()

    cur = conn.cursor()

    cur.execute(
        """
        INSERT INTO transactions(

            group_name,
            user,
            amount,
            category,
            type,
            note,
            date

        )
        VALUES(?,?,?,?,?,?,?)
        """,
        (
            session["group_name"],
            session["user"],
            amount,
            category,
            transaction_type,
            note,
            date.today().isoformat(),
        ),
    )

    conn.commit()

    conn.close()

    return redirect(url_for("dashboard"))


# =========================================================
# DELETE TRANSACTION
# =========================================================

@app.route("/delete/<int:id>")
def delete_transaction(id):

    if "user" not in session:

        return redirect(url_for("login"))

    conn = get_db()

    cur = conn.cursor()

    cur.execute(
        """
        DELETE FROM transactions
        WHERE id=?
        """,
        (id,)
    )

    conn.commit()

    conn.close()

    return redirect(url_for("dashboard"))


# =========================================================
# EDIT TRANSACTION
# =========================================================

@app.route("/edit/<int:id>", methods=["GET", "POST"])
def edit_transaction(id):

    if "user" not in session:

        return redirect(url_for("login"))

    conn = get_db()

    cur = conn.cursor()

    cur.execute(
        """
        SELECT * FROM transactions
        WHERE id=?
        """,
        (id,)
    )

    transaction = cur.fetchone()

    if not transaction:

        conn.close()

        return """
        <h2>Transaction Not Found</h2>
        """

    if request.method == "POST":

        amount = request.form["amount"]

        if not valid_amount(amount):

            conn.close()

            return """
            <h2>Invalid Amount</h2>

            <a href='/dashboard'>Go Back</a>
            """

        amount = float(amount)

        category = request.form["category"]
        if category == "Other":
            category = request.form.get("category_other", "").strip() or category

        transaction_type = request.form["type"]

        note = request.form["note"].strip()

        cur.execute(
            """
            UPDATE transactions

            SET
                amount=?,
                category=?,
                type=?,
                note=?

            WHERE id=?
            """,
            (
                amount,
                category,
                transaction_type,
                note,
                id,
            ),
        )

        conn.commit()

        conn.close()

        return redirect(url_for("dashboard"))

    conn.close()

    return render_template(
        "edit_transaction.html",
        transaction=transaction,
        role=session.get("role", "household")
    )


# =========================================================
# RUN
# =========================================================

if __name__ == "__main__":

    app.run(debug=True)
